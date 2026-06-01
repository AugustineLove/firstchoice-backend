from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

DARK_NAVY   = "0D1B2A"
MID_NAVY    = "1B2A3B"
ACCENT_GOLD = "C9A84C"
ACCENT_TEAL = "1ABC9C"
ACCENT_RED  = "E74C3C"
ACCENT_BLUE = "3498DB"
ACCENT_PURPLE = "9B59B6"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F4F6F9"
MID_GRAY    = "BDC3C7"
DARK_GRAY   = "7F8C8D"
INPUT_YELLOW = "FFFDE7"
INPUT_BORDER = "F9A825"

def F(hex_color): return PatternFill("solid", start_color=hex_color, fgColor=hex_color)
def FT(bold=False, size=11, color=WHITE, italic=False): return Font(bold=bold, size=size, color=color, italic=italic, name="Arial")
def AL(h="center", v="center", wrap=False): return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def S(color="CCCCCC", style="thin"): return Side(border_style=style, color=color)
def B_all(c="CCCCCC", s="thin"): ss=Side(s,color=c); return Border(left=ss,right=ss,top=ss,bottom=ss)
def B_left_thick(c,bg="E0E0E0",s="thin"): bg_s=Side(s,color=bg); return Border(left=Side("thick",color=c),right=bg_s,top=bg_s,bottom=bg_s)

def mc(ws, r, c, val=None, bold=False, size=11, fg=WHITE, bg=None, ha="center", va="center", wrap=False, italic=False, bdr=None, nf=None):
    cell = ws.cell(r, c, val)
    cell.font = Font(bold=bold, size=size, color=fg, italic=italic, name="Arial")
    if bg: cell.fill = F(bg)
    cell.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)
    if bdr: cell.border = bdr
    if nf: cell.number_format = nf
    return cell

ws = wb.active
ws.title = "Executive Dashboard"
ws.sheet_view.showGridLines = False
ws.sheet_view.zoomScale = 90

cw = {1:2,2:18,3:16,4:16,5:16,6:16,7:16,8:16,9:16,10:18,11:2}
for c,w in cw.items(): ws.column_dimensions[get_column_letter(c)].width = w
for r in range(1,75): ws.row_dimensions[r].height = 18
ws.row_dimensions[1].height = 8
ws.row_dimensions[2].height = 55
ws.row_dimensions[3].height = 12

for r in range(1,75):
    for c in range(1,12):
        ws.cell(r,c).fill = F(LIGHT_GRAY)

# HEADER
ws.merge_cells("B2:J2")
mc(ws,2,2,"EXECUTIVE PERFORMANCE DASHBOARD",True,22,WHITE,DARK_NAVY,"center","center")
ws.merge_cells("B3:J3")
mc(ws,3,2,"CONFIDENTIAL  •  FOR CEO REVIEW  •  FY 2025",False,9,ACCENT_GOLD,MID_NAVY,"center","center",italic=True)

def section_row(row, text):
    ws.row_dimensions[row].height = 22
    ws.merge_cells(f"B{row}:J{row}")
    mc(ws,row,2,f"  {text}",True,9,ACCENT_GOLD,MID_NAVY,"left","center")

# KPI CARD: draws a 4-row card. Row layout: title | big-value | sublabel | input cell
# top_row = first row of card, sc = start_col, nc = num_cols
def kpi_card(top_row, sc, nc, title, input_formula, accent):
    ec = sc + nc - 1
    for r in range(top_row, top_row+4):
        for c in range(sc, ec+1):
            ws.cell(r,c).fill = F(WHITE)
            ws.cell(r,c).border = B_all("E0E0E0")
        # left thick accent border on first col only
        ws.cell(r,sc).border = B_left_thick(accent)

    ws.merge_cells(f"{get_column_letter(sc)}{top_row}:{get_column_letter(ec)}{top_row}")
    mc(ws,top_row,sc,f"  {title}",True,8,DARK_GRAY,WHITE,"left","center")

    ws.merge_cells(f"{get_column_letter(sc)}{top_row+1}:{get_column_letter(ec)}{top_row+1}")
    val_cell = mc(ws,top_row+1,sc,input_formula,True,20,DARK_NAVY,WHITE,"center","center",nf="#,##0")

    ws.merge_cells(f"{get_column_letter(sc)}{top_row+2}:{get_column_letter(ec)}{top_row+2}")
    mc(ws,top_row+2,sc,"▼ Enter value below",False,7,DARK_GRAY,WHITE,"center","center",italic=True)

    ws.merge_cells(f"{get_column_letter(sc)}{top_row+3}:{get_column_letter(ec)}{top_row+3}")
    inp = ws.cell(top_row+3, sc)
    inp.fill = F(INPUT_YELLOW)
    inp.border = Border(left=Side("thick",color=accent),right=Side("thin",color=INPUT_BORDER),
                        top=Side("dashed",color=INPUT_BORDER),bottom=Side("medium",color=INPUT_BORDER))
    inp.alignment = AL()
    inp.font = Font(bold=True, size=13, color=DARK_NAVY, name="Arial")
    inp.number_format = "#,##0"
    return inp

def auto_card(top_row, sc, nc, title, display_formula, accent, pct=False):
    ec = sc + nc - 1
    for r in range(top_row, top_row+4):
        for c in range(sc, ec+1):
            ws.cell(r,c).fill = F(WHITE)
            ws.cell(r,c).border = B_all("E0E0E0")
        ws.cell(r,sc).border = B_left_thick(accent)

    ws.merge_cells(f"{get_column_letter(sc)}{top_row}:{get_column_letter(ec)}{top_row}")
    mc(ws,top_row,sc,f"  {title}",True,8,DARK_GRAY,WHITE,"left","center")

    ws.merge_cells(f"{get_column_letter(sc)}{top_row+1}:{get_column_letter(ec)}{top_row+1}")
    nf = "0.0%" if pct else "#,##0"
    mc(ws,top_row+1,sc,display_formula,True,20,DARK_NAVY,WHITE,"center","center",nf=nf)

    ws.merge_cells(f"{get_column_letter(sc)}{top_row+2}:{get_column_letter(ec)}{top_row+2}")
    mc(ws,top_row+2,sc,"Auto-Calculated  ✓",False,7,ACCENT_TEAL,WHITE,"center","center",italic=True)

    ws.merge_cells(f"{get_column_letter(sc)}{top_row+3}:{get_column_letter(ec)}{top_row+3}")
    auto = ws.cell(top_row+3, sc)
    auto.fill = F("E8F8F5")
    auto.border = Border(left=Side("thick",color=accent),right=Side("thin",color="1ABC9C"),
                         top=Side("dashed",color="1ABC9C"),bottom=Side("medium",color="1ABC9C"))
    auto.alignment = AL()
    auto.font = Font(bold=True, size=13, color=accent, name="Arial")
    auto.value = display_formula
    auto.number_format = nf

for r in range(1,75):
    for c in range(1,12):
        if not ws.cell(r,c).fill or ws.cell(r,c).fill.fgColor.rgb == "00000000":
            ws.cell(r,c).fill = F(LIGHT_GRAY)

# ── SEC 1: CUSTOMER METRICS ─────────────────────────────
section_row(5, "01  |  CUSTOMER METRICS")
for r in range(5,75):
    for c in range(1,12):
        if ws.cell(r,c).fill.fgColor.rgb == "00000000":
            ws.cell(r,c).fill = F(LIGHT_GRAY)

kpi_card(6, 2, 2, "ACTIVE CUSTOMERS", None, ACCENT_TEAL)
kpi_card(6, 4, 2, "INACTIVE CUSTOMERS", None, ACCENT_RED)
auto_card(6, 6, 2, "TOTAL CUSTOMERS", "=IF(C9+E9=0,\"\",C9+E9)", ACCENT_BLUE)
auto_card(6, 8, 2, "RETENTION RATE", '=IF(C9+E9=0,"",C9/(C9+E9))', ACCENT_GOLD, pct=True)

# ── SEC 2: REVENUE METRICS ──────────────────────────────
section_row(11, "02  |  REVENUE METRICS")
kpi_card(12, 2, 2, "TOTAL REVENUE ($)", None, ACCENT_GOLD)
kpi_card(12, 4, 2, "COST OF GOODS SOLD ($)", None, ACCENT_RED)
auto_card(12, 6, 2, "GROSS PROFIT ($)", "=IF(C15+E15=0,\"\",C15-E15)", ACCENT_TEAL)
auto_card(12, 8, 2, "GROSS MARGIN (%)", '=IF(C15=0,"",((C15-E15)/C15))', ACCENT_PURPLE, pct=True)

# ── SEC 3: OPERATIONS ───────────────────────────────────
section_row(17, "03  |  OPERATIONAL METRICS")
kpi_card(18, 2, 2, "NEW LEADS", None, ACCENT_BLUE)
kpi_card(18, 4, 2, "CONVERSIONS", None, ACCENT_TEAL)
auto_card(18, 6, 2, "CONVERSION RATE (%)", '=IF(C21=0,"",E21/C21)', ACCENT_GOLD, pct=True)
kpi_card(18, 8, 2, "OPEN SUPPORT TICKETS", None, ACCENT_RED)

# ── SEC 4: FINANCIAL ────────────────────────────────────
section_row(23, "04  |  FINANCIAL HEALTH")
kpi_card(24, 2, 2, "OPERATING EXPENSES ($)", None, ACCENT_RED)
kpi_card(24, 4, 2, "NET PROFIT ($)", None, ACCENT_TEAL)
auto_card(24, 6, 2, "NET PROFIT MARGIN (%)", '=IF(C15=0,"",E27/(C15))', ACCENT_PURPLE, pct=True)
kpi_card(24, 8, 2, "CASH BALANCE ($)", None, ACCENT_GOLD)

# ── SEC 5: MONTHLY TABLE ────────────────────────────────
section_row(29, "05  |  MONTHLY PERFORMANCE TRACKER")
ws.row_dimensions[30].height = 24

# fix background for all used area
for r in range(1,80):
    for c in range(1,12):
        cell = ws.cell(r,c)
        if cell.fill.fgColor.type == 'rgb' and cell.fill.fgColor.rgb == "00000000":
            cell.fill = F(LIGHT_GRAY)

hdrs = ["MONTH","ACTIVE","INACTIVE","TOTAL","REVENUE ($)","EXPENSES ($)","NET PROFIT ($)","MARGIN %"]
for i, h in enumerate(hdrs):
    c = ws.cell(30, i+2)
    c.value = h; c.font = Font(bold=True, size=8, color=WHITE, name="Arial")
    c.fill = F(DARK_NAVY); c.alignment = AL()
    c.border = Border(bottom=Side("medium",color=ACCENT_GOLD), left=Side("thin",color=MID_NAVY), right=Side("thin",color=MID_NAVY))

months=["January","February","March","April","May","June","July","August","September","October","November","December"]
for i, mo in enumerate(months):
    r = 31 + i
    ws.row_dimensions[r].height = 20
    bg = WHITE if i%2==0 else "EEF2F7"

    # Month
    c=ws.cell(r,2); c.value=mo; c.font=Font(bold=True,size=9,color=DARK_NAVY,name="Arial"); c.fill=F(bg); c.alignment=AL("left"); c.border=Border(left=Side("medium",color=DARK_NAVY),right=Side("thin",color="E0E0E0"),top=Side("thin",color="E0E0E0"),bottom=Side("thin",color="E0E0E0"))
    # Active - input
    for col in [3,4,6,7]:
        cc=ws.cell(r,col); cc.fill=F(INPUT_YELLOW); cc.border=B_all(INPUT_BORDER,"thin"); cc.alignment=AL(); cc.font=Font(size=9,color=DARK_NAVY,name="Arial"); cc.number_format="#,##0"
    # Total auto
    tc=ws.cell(r,5); tc.value=f"=IF(C{r}+D{r}=0,\"\",C{r}+D{r})"; tc.fill=F("E8F8F5"); tc.border=B_all("1ABC9C","thin"); tc.alignment=AL(); tc.font=Font(bold=True,size=9,color=ACCENT_TEAL,name="Arial"); tc.number_format="#,##0"
    # Net Profit auto
    np_=ws.cell(r,8); np_.value=f"=IF(F{r}+G{r}=0,\"\",F{r}-G{r})"; np_.fill=F("E8F8F5"); np_.border=B_all("1ABC9C","thin"); np_.alignment=AL(); np_.font=Font(bold=True,size=9,color=DARK_NAVY,name="Arial"); np_.number_format="#,##0"
    # Margin auto
    mg=ws.cell(r,9); mg.value=f'=IF(F{r}=0,"",H{r}/F{r})'; mg.fill=F("F3E5F5"); mg.border=Border(right=Side("medium",color=DARK_NAVY),left=Side("thin",color="E0E0E0"),top=Side("thin",color="E0E0E0"),bottom=Side("thin",color="E0E0E0")); mg.alignment=AL(); mg.font=Font(bold=True,size=9,color=ACCENT_PURPLE,name="Arial"); mg.number_format="0.0%"

# Totals
r_t=43; ws.row_dimensions[r_t].height=22
ws.cell(r_t,2).value="ANNUAL TOTALS"; ws.cell(r_t,2).font=Font(bold=True,size=9,color=ACCENT_GOLD,name="Arial"); ws.cell(r_t,2).fill=F(MID_NAVY); ws.cell(r_t,2).alignment=AL("left")
for ci,cl,f_ in [(3,"C","#,##0"),(4,"D","#,##0"),(5,"E","#,##0"),(6,"F","#,##0"),(7,"G","#,##0"),(8,"H","#,##0")]:
    cc=ws.cell(r_t,ci); cc.value=f"=SUM({cl}31:{cl}42)"; cc.font=Font(bold=True,size=9,color=WHITE,name="Arial"); cc.fill=F(MID_NAVY); cc.alignment=AL(); cc.number_format=f_; cc.border=B_all(DARK_NAVY)
mg_=ws.cell(r_t,9); mg_.value='=IF(F43=0,"",H43/F43)'; mg_.font=Font(bold=True,size=9,color=ACCENT_GOLD,name="Arial"); mg_.fill=F(MID_NAVY); mg_.alignment=AL(); mg_.number_format="0.0%"; mg_.border=B_all(DARK_NAVY)
ws.cell(r_t,2).border=Border(left=Side("medium",color=DARK_NAVY),right=Side("thin",color=DARK_NAVY),top=Side("thin",color=DARK_NAVY),bottom=Side("thin",color=DARK_NAVY))

# Legend
ws.row_dimensions[45].height = 8
ws.row_dimensions[46].height = 22
ws.merge_cells("B46:J46")
leg=ws.cell(46,2); leg.value="  🟡 Yellow = Enter your figures     🟢 Green = Auto-calculated     🟣 Purple = Auto-calculated %     |  Active + Inactive = Total  |  Revenue – COGS = Gross Profit  |  Net Profit ÷ Revenue = Margin"
leg.font=Font(size=8,color=DARK_GRAY,italic=True,name="Arial"); leg.fill=F(WHITE); leg.alignment=AL("left")
leg.border=Border(top=Side("medium",color=MID_GRAY),bottom=Side("medium",color=MID_GRAY))

# Footer
ws.row_dimensions[48].height = 8
ws.row_dimensions[49].height = 22
ws.merge_cells("B49:J49")
foot=ws.cell(49,2); foot.value="CONFIDENTIAL  —  This document is prepared exclusively for executive review. Distribution requires written authorization."
foot.font=Font(size=8,color=WHITE,italic=True,name="Arial"); foot.fill=F(DARK_NAVY); foot.alignment=AL("center")

# ── ENSURE FULL BACKGROUND ─────────────────────────────
for r in range(1,55):
    for c in range(1,12):
        cell = ws.cell(r,c)
        try:
            if cell.fill.fgColor.type == 'rgb' and cell.fill.fgColor.rgb in ("00000000",""):
                cell.fill = F(LIGHT_GRAY)
        except: pass

# ══════════════════════════════════════════════════════
# SHEET 2: INPUT GUIDE
# ══════════════════════════════════════════════════════
ws2 = wb.create_sheet("Input Reference Guide")
ws2.sheet_view.showGridLines = False
ws2.column_dimensions["A"].width = 2
ws2.column_dimensions["B"].width = 35
ws2.column_dimensions["C"].width = 22
ws2.column_dimensions["D"].width = 50

for r in range(1,45):
    for c in range(1,6): ws2.cell(r,c).fill = F(LIGHT_GRAY)

ws2.row_dimensions[1].height = 45
ws2.merge_cells("B1:D1")
mc(ws2,1,2,"DATA INPUT REFERENCE GUIDE",True,18,WHITE,DARK_NAVY,"center","center")

ws2.row_dimensions[2].height = 8
ws2.row_dimensions[3].height = 22
for i,h in enumerate(["METRIC","DASHBOARD CELL","DESCRIPTION"]):
    cc=ws2.cell(3,i+2); cc.value=h; cc.font=Font(bold=True,size=9,color=WHITE,name="Arial"); cc.fill=F(MID_NAVY); cc.alignment=AL(); cc.border=Border(bottom=Side("medium",color=ACCENT_GOLD))

guide=[
    ("Active Customers",       "Row 6–9, Col B–C",  "Enter count of currently active customers",        INPUT_YELLOW, DARK_NAVY, False),
    ("Inactive Customers",     "Row 6–9, Col D–E",  "Enter count of churned / dormant customers",       INPUT_YELLOW, DARK_NAVY, False),
    ("Total Customers",        "Row 6–9, Col F–G",  "AUTO — Active + Inactive",                         "E8F8F5",    ACCENT_TEAL, True),
    ("Retention Rate",         "Row 6–9, Col H–I",  "AUTO — Active ÷ Total Customers",                  "E8F8F5",    ACCENT_TEAL, True),
    ("Total Revenue ($)",      "Row 12–15, Col B–C","Enter gross revenue for the period",               INPUT_YELLOW, DARK_NAVY, False),
    ("Cost of Goods Sold ($)", "Row 12–15, Col D–E","Enter direct costs attributable to revenue",       INPUT_YELLOW, DARK_NAVY, False),
    ("Gross Profit ($)",       "Row 12–15, Col F–G","AUTO — Revenue − COGS",                            "E8F8F5",    ACCENT_TEAL, True),
    ("Gross Margin (%)",       "Row 12–15, Col H–I","AUTO — Gross Profit ÷ Revenue",                    "F3E5F5",    ACCENT_PURPLE, True),
    ("New Leads",              "Row 18–21, Col B–C","Total new leads generated in the period",          INPUT_YELLOW, DARK_NAVY, False),
    ("Conversions",            "Row 18–21, Col D–E","Leads that converted to paying customers",         INPUT_YELLOW, DARK_NAVY, False),
    ("Conversion Rate (%)",    "Row 18–21, Col F–G","AUTO — Conversions ÷ New Leads",                   "E8F8F5",    ACCENT_GOLD, True),
    ("Open Support Tickets",   "Row 18–21, Col H–I","Count of unresolved customer support tickets",     INPUT_YELLOW, DARK_NAVY, False),
    ("Operating Expenses ($)", "Row 24–27, Col B–C","Total operating costs excluding COGS",             INPUT_YELLOW, DARK_NAVY, False),
    ("Net Profit ($)",         "Row 24–27, Col D–E","Enter net profit after all costs",                 INPUT_YELLOW, DARK_NAVY, False),
    ("Net Profit Margin (%)",  "Row 24–27, Col F–G","AUTO — Net Profit ÷ Revenue",                      "F3E5F5",    ACCENT_PURPLE, True),
    ("Cash Balance ($)",       "Row 24–27, Col H–I","Current cash / total bank balance",                INPUT_YELLOW, DARK_NAVY, False),
]
for i,(metric,cell_ref,desc,bg,fg,is_auto) in enumerate(guide):
    r=4+i; ws2.row_dimensions[r].height=19
    bg2 = WHITE if i%2==0 else "F7F9FC"
    m=ws2.cell(r,2); m.value=("🔒 " if is_auto else "✏️  ")+metric; m.font=Font(size=9,color=fg if is_auto else DARK_NAVY,name="Arial"); m.fill=F(bg); m.alignment=AL("left"); m.border=B_all("E0E0E0")
    cr=ws2.cell(r,3); cr.value=cell_ref; cr.font=Font(bold=True,size=9,color=fg if is_auto else DARK_NAVY,name="Arial"); cr.fill=F(bg); cr.alignment=AL("center"); cr.border=B_all("E0E0E0")
    d=ws2.cell(r,4); d.value=desc; d.font=Font(size=9,color=DARK_GRAY,italic=is_auto,name="Arial"); d.fill=F(bg); d.alignment=AL("left"); d.border=B_all("E0E0E0")

ws2.row_dimensions[21].height = 12
ws2.merge_cells("B22:D22")
n=ws2.cell(22,2); n.value="  ✏️  = Manual input required     🔒 = Auto-calculated, do not edit"
n.font=Font(size=9,color=DARK_GRAY,italic=True,name="Arial"); n.fill=F(WHITE); n.alignment=AL("left")
n.border=Border(top=Side("medium",color=MID_GRAY),bottom=Side("medium",color=MID_GRAY))

wb.save("CEO_Executive_Dashboard.xlsx")
print("Done")